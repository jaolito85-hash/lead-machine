"""
LEAD MACHINE — Export de campanha pra entrega ao cliente.

Gera CSV ou XLSX padronizado contendo todos os leads de uma campanha.
Nome de arquivo identifica cliente_destino + campanha + data + total.

Uso programatico:
    from agents import exports
    path = exports.export_campaign("C-0001", "xlsx")

Uso via endpoint (serve.py):
    GET /api/local/exports?campaign_id=C-0001&format=xlsx
"""

import csv
import io
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

import campaigns as campaigns_module
from base import LEADS_DIR, load_leads

EXPORTS_DIR = LEADS_DIR / "exports"

VALID_FORMATS = {"csv", "xlsx"}

# Ordem de severidade pra ordenar leads (quente primeiro).
TEMP_RANK = {"quente": 0, "morno": 1, "frio": 2, "": 3}

# Schema unico (pessoa + empresa). Tupla (titulo, chave_no_lead).
EXPORT_COLUMNS = [
    ("ID",              "id"),
    ("Tipo",            "tipo"),
    ("Nome",            "nome"),
    ("Plataforma",      "plataforma"),
    ("Perfil",          "perfil"),
    ("URL do Perfil",   "author_url"),
    ("Cidade",          "cidade"),
    ("Endereco",        "endereco"),
    ("Nicho",           "nicho"),
    ("Temperatura",     "temp"),
    ("Score",           "score"),
    ("Evidencia",       "evidencia"),
    ("Texto original",  "texto_original"),
    ("URL do Post",     "url"),
    ("Post owner",      "post_owner"),
    ("Intent",          "intent"),
    ("Urgencia",        "urgency"),
    ("Website",         "website"),
    ("Rating",          "rating"),
    ("Email",           "email"),
    ("Telefone",        "telefone"),
    ("Coletado em",     "coletado"),
]


# ────────────────────────────────────────────────────────────
# Utils
# ────────────────────────────────────────────────────────────

def slugify(text: str, max_len: int = 60) -> str:
    """Converte texto livre num slug seguro pra nome de arquivo."""
    if not text:
        return ""
    # Normaliza acentos: 'Câmpo Mourão' -> 'Campo Mourao'
    norm = unicodedata.normalize("NFKD", text)
    ascii_only = "".join(c for c in norm if not unicodedata.combining(c))
    # Lower, troca nao-alfanumerico por hifen, colapsa hifens
    s = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_only).strip("-").lower()
    return s[:max_len] if s else ""


def _ordering_key(lead: dict):
    """Ordena por temp (quente -> morno -> frio), score desc, scraped_at desc."""
    return (
        TEMP_RANK.get(lead.get("temp", ""), 3),
        -int(lead.get("score") or 0),
        -_iso_to_ts(lead.get("scraped_at", "")),
    )


def _iso_to_ts(iso: str) -> float:
    if not iso:
        return 0.0
    try:
        return datetime.fromisoformat(iso).timestamp()
    except (ValueError, TypeError):
        return 0.0


def _value_for_column(lead: dict, key: str):
    """Le valor da coluna do lead, normalizando None e tipos compostos."""
    v = lead.get(key)
    if v is None:
        return ""
    return v


def _filename_for(campaign: dict, count: int, fmt: str) -> str:
    cliente_slug = slugify(campaign.get("cliente_destino", "")) or "sem-cliente"
    camp_slug = slugify(campaign.get("nome", "")) or campaign.get("id", "campanha").lower()
    today = datetime.now().strftime("%Y-%m-%d")
    return f"{cliente_slug}_{camp_slug}_{today}_{count}leads.{fmt}"


# ────────────────────────────────────────────────────────────
# Renderers
# ────────────────────────────────────────────────────────────

def _render_csv(campaign: dict, leads: list) -> bytes:
    """CSV com BOM utf-8-sig pra Excel BR abrir com acento certo."""
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
    writer.writerow([title for title, _ in EXPORT_COLUMNS])
    for lead in leads:
        writer.writerow([_value_for_column(lead, key) for _, key in EXPORT_COLUMNS])
    return buf.getvalue().encode("utf-8-sig")


def _render_xlsx(campaign: dict, leads: list) -> bytes:
    """XLSX com cabecalho de campanha + tabela. Requer openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Bloco de metadados (4 linhas)
    cliente = campaign.get("cliente_destino") or "(sem cliente)"
    nicho = campaign.get("nicho") or "-"
    cidade = campaign.get("cidade") or ("Nacional" if campaign.get("nacional") else "-")
    today = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws["A1"] = f"Cliente: {cliente}"
    ws["A2"] = f"Campanha: {campaign.get('nome', '')} ({campaign.get('id', '')})"
    ws["A3"] = f"Nicho: {nicho}  |  Cidade: {cidade}"
    ws["A4"] = f"Total de leads: {len(leads)}  |  Gerado em: {today}"
    for row in (1, 2, 3, 4):
        ws.cell(row=row, column=1).font = bold

    header_row = 6
    for col_idx, (title, _) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_offset, lead in enumerate(leads, start=1):
        for col_idx, (_, key) in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=header_row + r_offset, column=col_idx,
                           value=_value_for_column(lead, key))
            cell.alignment = wrap

    # Largura de coluna heuristica (limitada).
    widths = {}
    for col_idx, (title, key) in enumerate(EXPORT_COLUMNS, start=1):
        max_len = len(title)
        for lead in leads:
            v = _value_for_column(lead, key)
            l = len(str(v)) if v != "" else 0
            if l > max_len:
                max_len = l
        widths[col_idx] = min(max(max_len + 2, 10), 60)
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ────────────────────────────────────────────────────────────
# Public API
# ────────────────────────────────────────────────────────────

class ExportError(Exception):
    """Erro de export. Tem .code: 'campaign_not_found', 'invalid_format', 'empty'."""

    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def build_export(campaign_id: str, fmt: str = "csv") -> tuple:
    """
    Gera o export em memoria. Nao escreve em disco — usado pelo endpoint HTTP.

    Retorna (filename, content_bytes, lead_count, campaign_dict).

    Lanca ExportError se campanha nao existe ou formato invalido.
    Se nao houver leads, retorna content_bytes vazio com lead_count=0
    (chamador decide o que fazer — endpoint retorna 200 com header X-Lead-Count: 0).
    """
    fmt = (fmt or "csv").lower()
    if fmt not in VALID_FORMATS:
        raise ExportError("invalid_format", f"formato invalido: {fmt} (use csv ou xlsx)")

    campaign = campaigns_module.get(campaign_id)
    if campaign is None:
        raise ExportError("campaign_not_found", f"campanha nao encontrada: {campaign_id}")

    all_leads = load_leads()
    leads = [l for l in all_leads if l.get("campaign_id") == campaign_id]
    leads.sort(key=_ordering_key)

    filename = _filename_for(campaign, len(leads), fmt)

    if not leads:
        return (filename, b"", 0, campaign)

    if fmt == "csv":
        content = _render_csv(campaign, leads)
    else:
        content = _render_xlsx(campaign, leads)

    return (filename, content, len(leads), campaign)


def export_campaign(campaign_id: str, fmt: str = "csv") -> Optional[Path]:
    """
    Gera arquivo de export e SALVA em leads-export/exports/.
    Atualiza campaigns.json com last_exported_at + last_exported_count.

    Retorna Path do arquivo escrito, ou None se a campanha nao tiver leads.
    """
    filename, content, count, _ = build_export(campaign_id, fmt)

    if count == 0:
        return None

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORTS_DIR / filename
    path.write_bytes(content)

    campaigns_module.update(campaign_id, {
        "last_exported_at": datetime.now().isoformat(timespec="seconds"),
        "last_exported_count": count,
        "last_exported_format": fmt,
    })
    return path


# ────────────────────────────────────────────────────────────
# CLI util (debug)
# ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("uso: python -m agents.exports <campaign_id> [csv|xlsx]")
        sys.exit(1)
    cid = sys.argv[1]
    fmt = sys.argv[2] if len(sys.argv) >= 3 else "csv"
    p = export_campaign(cid, fmt)
    if p is None:
        print(f"campanha {cid}: 0 leads — nao gerou arquivo")
    else:
        print(f"OK: {p}")
